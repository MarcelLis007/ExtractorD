import openpyxl
import pandas as pd
import os
import re
from pathlib import Path
from datetime import datetime

class ExtractorFormulariosCompleto:
    """
    Extractor optimizado basado en análisis del documento real.
    """
    
    def __init__(self, carpeta_excel):
        self.carpeta = Path(carpeta_excel)
    
    def limpiar_texto(self, texto):
        """Limpia texto eliminando espacios y valores nulos."""
        if texto is None:
            return None
        texto = str(texto).strip()
        # Eliminar valores que claramente no son datos
        if texto in ['', 'null', 'NULL', 'None', '-', 'N/A', 'XXXXXXX', 'XXXXXX', 
                     'd', '_', '...', 'null', 'SIN VALOR']:
            return None
        return texto
    
    def tiene_fondo_amarillo(self, celda):
        """Detecta si una celda tiene fondo amarillo."""
        try:
            if celda.fill and celda.fill.start_color:
                color = celda.fill.start_color
                if color.type == 'rgb':
                    rgb = str(color.rgb).upper()
                    # Amarillo: FFFFFF00, FFFF00, FFFFCC, etc.
                    if 'FFFF' in rgb or ('FF' in rgb[:4] and 'FF' in rgb[2:6]):
                        return True
                elif color.index and color.index in [13, 43, 65]:  # Índices de amarillo en Excel
                    return True
        except:
            pass
        return False
    
    def buscar_con_amarillo(self, sheet, etiquetas, max_fila=25):
        """
        Busca celdas con fondo amarillo cerca de una etiqueta.
        Usado para CALIFICACIÓN y REVISADO.
        """
        for fila in range(1, min(max_fila, sheet.max_row + 1)):
            for col in range(1, sheet.max_column + 1):
                celda = sheet.cell(fila, col)
                texto = str(celda.value or '').strip().upper()
                
                # Verificar si es la etiqueta buscada
                for etiqueta in etiquetas:
                    if etiqueta.upper() in texto:
                        # Buscar celdas amarillas en un rango amplio
                        for f in range(max(1, fila-1), min(fila+3, sheet.max_row+1)):
                            for c in range(col, min(col+10, sheet.max_column+1)):
                                celda_check = sheet.cell(f, c)
                                if self.tiene_fondo_amarillo(celda_check):
                                    valor = self.limpiar_texto(celda_check.value)
                                    if valor and len(valor) <= 30:  # Valores cortos
                                        return valor
        return None
    
    def buscar_valor_simple(self, sheet, etiquetas, max_fila=60, tipo_dato='texto'):
        """
        Busca un valor simple cerca de una etiqueta.
        Retorna SOLO el primer valor válido encontrado.
        
        tipo_dato puede ser: 'texto', 'numero', 'fecha', 'cedula', 'alfanumerico'
        """
        for fila in range(1, min(max_fila, sheet.max_row + 1)):
            for col in range(1, sheet.max_column + 1):
                texto_celda = str(sheet.cell(fila, col).value or '').strip()
                
                if not texto_celda:
                    continue
                
                # Verificar coincidencia con etiqueta
                for etiqueta in etiquetas:
                    if etiqueta.upper() in texto_celda.upper():
                        # Caso 1: Valor en la misma celda después de ":"
                        if ':' in texto_celda and len(texto_celda) > len(etiqueta) + 2:
                            valor = texto_celda.split(':', 1)[1].strip()
                            valor_limpio = self.limpiar_texto(valor)
                            if valor_limpio and self._validar_tipo_dato(valor_limpio, tipo_dato):
                                return valor_limpio
                        
                        # Caso 2: Buscar en celdas adyacentes
                        posiciones = [
                            (fila, col+1), (fila, col+2), (fila, col+3),
                            (fila+1, col), (fila+1, col+1), (fila-1, col+1)
                        ]
                        
                        for f, c in posiciones:
                            if 1 <= f <= sheet.max_row and 1 <= c <= sheet.max_column:
                                candidato = self.limpiar_texto(sheet.cell(f, c).value)
                                if candidato and len(candidato) > 0:
                                    # Verificar que no sea otra etiqueta
                                    es_etiqueta = any(x in candidato.upper() for x in [
                                        'CI:', 'SCORE', 'BANCO', 'CUENTA', 'FECHA:', 
                                        'CIFRAS', 'CHP', 'APERTURA', 'AÑO', 'CUPO', 
                                        'COMENTARIO', 'TOTAL', 'EMPRESA', 'GARANTIA:',
                                        'GARANTE:', 'TITULAR', 'CONYUGUE', 'CÓNYUGE'
                                    ])
                                    if not es_etiqueta and self._validar_tipo_dato(candidato, tipo_dato):
                                        return candidato
        return None
    
    def _validar_tipo_dato(self, valor, tipo_dato):
        """Valida que el valor corresponda al tipo de dato esperado."""
        if tipo_dato == 'numero':
            # Debe contener números, puede tener $, comas, puntos
            return bool(re.search(r'\d', valor))
        elif tipo_dato == 'fecha':
            # Formato de fecha: dd/mm/yyyy o similar
            return bool(re.search(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', valor))
        elif tipo_dato == 'cedula':
            # Cédula: 10 o 13 dígitos
            return bool(re.search(r'\d{10,13}', valor))
        elif tipo_dato == 'alfanumerico':
            # Letras y números mezclados (para matrícula)
            tiene_letra = bool(re.search(r'[A-Za-z]', valor))
            tiene_numero = bool(re.search(r'\d', valor))
            return tiene_letra and tiene_numero
        else:  # texto
            return True
    
    def extraer_ci_garante(self, sheet):
        """
        Extrae CI del GARANTE (número de cédula 10 dígitos).
        NO debe confundirse con fechas o descripciones.
        """
        for fila in range(1, 30):
            for col in range(1, sheet.max_column + 1):
                texto = str(sheet.cell(fila, col).value or '').strip().upper()
                
                if 'CI: GARANTE' in texto or 'CI GARANTE' in texto:
                    # Buscar número de cédula (10 dígitos)
                    for c in range(col, min(col+5, sheet.max_column+1)):
                        val = str(sheet.cell(fila, c).value or '').strip()
                        # Buscar exactamente 10 dígitos (no 13 como RUC)
                        match = re.search(r'\b(\d{10})\b', val)
                        if match:
                            return match.group(1)
        return None
    
    def extraer_score_garante(self, sheet):
        """
        Extrae SCORE GARANTE (número o descripción tipo score).
        NO debe confundir con GARANTIA:.
        """
        for fila in range(1, 40):
            for col in range(1, sheet.max_column + 1):
                texto = str(sheet.cell(fila, col).value or '').strip().upper()
                
                if texto == 'SCORE GARANTE':
                    # Buscar valor numérico o descripción de score
                    for c in range(col+1, min(col+6, sheet.max_column+1)):
                        val = str(sheet.cell(fila, c).value or '').strip()
                        if val and 'GARANTIA' not in val.upper():
                            # Debe tener números o palabras relacionadas con crédito
                            if re.search(r'\d', val) or any(x in val.upper() for x in ['PRESTAMO', 'CREDITO', 'DIA', 'ATRASO']):
                                return val
        return None
    
    def extraer_garante_si_no(self, sheet):
        """
        Extrae GARANTE: debe retornar SI o NO (o nombre del garante).
        """
        for fila in range(1, 40):
            for col in range(1, sheet.max_column + 1):
                texto = str(sheet.cell(fila, col).value or '').strip().upper()
                
                if texto == 'GARANTE:':
                    # Buscar valor
                    for c in range(col+1, min(col+4, sheet.max_column+1)):
                        val = str(sheet.cell(fila, c).value or '').strip()
                        if val:
                            val_upper = val.upper()
                            # Si es XXXXXX o vacío = NO
                            if 'XXXX' in val_upper or val in ['-', '_']:
                                return 'NO'
                            # Si tiene nombre
                            elif len(val) > 3 and not val.isdigit():
                                return val
                            # Si dice SI explícitamente
                            elif 'SI' in val_upper:
                                return 'SI'
                    return 'NO'  # Por defecto si no encuentra nada
        return None
    
    def extraer_cupo(self, sheet):
        """
        Extrae CUPO: debe ser un número (valor monetario).
        """
        for fila in range(1, 50):
            for col in range(1, sheet.max_column + 1):
                texto = str(sheet.cell(fila, col).value or '').strip().upper()
                
                if 'CUPO:' in texto:
                    # Buscar número
                    for c in range(col, min(col+4, sheet.max_column+1)):
                        val = str(sheet.cell(fila, c).value or '').strip()
                        # Extraer solo número
                        match = re.search(r'[\$]?\s*(\d+[\.,]?\d*)', val)
                        if match:
                            return match.group(0)
        return None
    
    def extraer_cliente_desde(self, sheet):
        """
        Extrae CLIENTE DESDE: debe ser una FECHA.
        """
        for fila in range(1, 50):
            for col in range(1, sheet.max_column + 1):
                texto = str(sheet.cell(fila, col).value or '').strip().upper()
                
                if 'CLIENTE DESDE' in texto:
                    # Buscar fecha en filas siguientes
                    for f in range(fila, min(fila+5, sheet.max_row+1)):
                        for c in range(col, min(col+5, sheet.max_column+1)):
                            val = str(sheet.cell(f, c).value or '').strip()
                            # Buscar formato de fecha
                            match = re.search(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', val)
                            if match:
                                return match.group(0)
        return None
    
    def extraer_matricula_vehiculo(self, sheet):
        """
        Extrae MATRICULA VEHICULO: alfanumérico (letras y números).
        Ejemplo: GSB-4512, ABC-123, etc.
        """
        for fila in range(1, 40):
            for col in range(1, sheet.max_column + 1):
                texto = str(sheet.cell(fila, col).value or '').strip().upper()
                
                if 'MATRICULA VEHICULO' in texto or 'MATRÍCULA VEHÍCULO' in texto:
                    # Buscar valor
                    for c in range(col+1, min(col+5, sheet.max_column+1)):
                        val = str(sheet.cell(fila, c).value or '').strip()
                        if val:
                            val_upper = val.upper()
                            # Si dice NO o SI
                            if val_upper in ['NO', 'SI']:
                                return val_upper
                            # Si es matrícula (letras y números)
                            elif re.search(r'[A-Z]{2,3}[-\s]?\d{3,4}', val_upper):
                                return val
                            # Si tiene descripción con matrícula
                            elif 'MATRICULA' in val_upper and re.search(r'\d{4}', val):
                                return val
        return None
    
    def extraer_por_vencer(self, sheet):
        """
        Extrae POR VENCER: debe ser un VALOR monetario.
        """
        for fila in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                texto = str(sheet.cell(fila, col).value or '').strip().upper()
                
                if 'POR VENCER' in texto:
                    # Buscar valor monetario
                    for c in range(col, min(col+4, sheet.max_column+1)):
                        val = str(sheet.cell(fila, c).value or '').strip()
                        # Debe tener números o símbolo $
                        match = re.search(r'[\$]?\s*(\d+[\.,]?\d*)', val)
                        if match and 'FECHA' not in val.upper() and 'CHP' not in val.upper():
                            return match.group(0)
        return None
    
    def extraer_codigo_unico(self, sheet):
        """Extrae código único de la esquina superior derecha."""
        # Buscar en las primeras 3 filas, últimas 5 columnas
        for fila in range(1, 4):
            for col in range(sheet.max_column, max(sheet.max_column - 5, 0), -1):
                valor = self.limpiar_texto(sheet.cell(fila, col).value)
                if valor:
                    # Buscar número de 5+ dígitos
                    match = re.search(r'(\d{5,})', valor)
                    if match:
                        return match.group(1)
        return None
    
    def extraer_calificacion(self, sheet):
        """
        Extrae calificación: busca A, B o C con fondo amarillo.
        """
        # Primero intentar con color amarillo
        calificacion = self.buscar_con_amarillo(sheet, ['CALIFICACIÓN', 'CALIFICACION'], max_fila=15)
        if calificacion and calificacion in ['A', 'B', 'C']:
            return calificacion
        
        # Si no encuentra con amarillo, buscar la letra sola
        for fila in range(1, 15):
            for col in range(1, sheet.max_column + 1):
                texto = str(sheet.cell(fila, col).value or '').strip().upper()
                if 'CALIFICACIÓN' in texto or 'CALIFICACION' in texto:
                    # Buscar A, B, C en celdas cercanas
                    for c in range(col, min(col+5, sheet.max_column+1)):
                        val = self.limpiar_texto(sheet.cell(fila, c).value)
                        if val in ['A', 'B', 'C']:
                            return val
        return None
    
    def extraer_carpeta_completa(self, sheet):
        """
        Extrae si tiene carpeta completa (busca amarillo en REVISADO).
        """
        revisado = self.buscar_con_amarillo(sheet, ['REVISADO'], max_fila=15)
        if revisado and 'COMPLETA' in revisado.upper():
            return 'SI'
        
        # Verificar sin amarillo
        for fila in range(1, 15):
            for col in range(sheet.max_column - 5, sheet.max_column + 1):
                if col >= 1:
                    texto = str(sheet.cell(fila, col).value or '').strip().upper()
                    if 'COMPLETA' in texto or 'COMPLETO' in texto:
                        return 'SI'
        return 'NO'
    
    def extraer_ruc_y_anio(self, sheet):
        """
        Extrae RUC (número 13 dígitos) y AÑO (20XX) por separado.
        """
        ruc = None
        anio = None
        
        for fila in range(1, 20):
            for col in range(1, sheet.max_column + 1):
                texto = str(sheet.cell(fila, col).value or '').strip().upper()
                
                if 'RUC' in texto or 'AÑO' in texto:
                    # Buscar en celdas siguientes
                    for c in range(col, min(col+6, sheet.max_column+1)):
                        val = str(sheet.cell(fila, c).value or '').strip()
                        
                        # Extraer RUC (13 dígitos)
                        if not ruc:
                            match_ruc = re.search(r'(\d{13})', val)
                            if match_ruc:
                                ruc = match_ruc.group(1)
                        
                        # Extraer AÑO (20XX)
                        if not anio:
                            match_anio = re.search(r'(20\d{2})', val)
                            if match_anio:
                                anio = match_anio.group(1)
                    
                    if ruc and anio:
                        return ruc, anio
        
        return ruc, anio
    
    def extraer_edad(self, sheet):
        """
        Extrae edad (número entre 18-100).
        Nota: A veces aparece duplicado "30 | 30".
        """
        for fila in range(1, 20):
            for col in range(1, sheet.max_column + 1):
                texto = str(sheet.cell(fila, col).value or '').strip().upper()
                
                if 'EDAD' in texto:
                    # Buscar número
                    for c in range(col, min(col+5, sheet.max_column+1)):
                        val = str(sheet.cell(fila, c).value or '').strip()
                        # Buscar número solo
                        match = re.search(r'\b(\d{2})\b', val)
                        if match:
                            edad = int(match.group(1))
                            if 18 <= edad <= 100:
                                return str(edad)
        return None
    
    def extraer_vendedor_ciudad(self, sheet):
        """
        Extrae VENDEDOR y CIUDAD que pueden estar juntos o separados.
        Formato: "VENDEDOR: Pato Cueva" en una celda, "CIUDAD: QUITO" en otra o juntos.
        """
        vendedor = None
        ciudad = None
        
        for fila in range(1, 50):
            for col in range(1, sheet.max_column + 1):
                texto = str(sheet.cell(fila, col).value or '').strip()
                
                # Buscar VENDEDOR
                if 'VENDEDOR:' in texto.upper() and not vendedor:
                    # Caso 1: Mismo texto tiene CIUDAD también
                    if 'CIUDAD:' in texto.upper():
                        partes = texto.split('CIUDAD:', 1)
                        vendedor_parte = partes[0].replace('VENDEDOR:', '').replace('Vendedor:', '').strip()
                        ciudad_parte = partes[1].strip() if len(partes) > 1 else None
                        
                        vendedor = self.limpiar_texto(vendedor_parte) if vendedor_parte else None
                        ciudad = self.limpiar_texto(ciudad_parte) if ciudad_parte else None
                    else:
                        # Solo VENDEDOR
                        vendedor = texto.split(':', 1)[1].strip() if ':' in texto else None
                        vendedor = self.limpiar_texto(vendedor)
                
                # Buscar CIUDAD si no se encontró antes
                if 'CIUDAD:' in texto.upper() and not ciudad:
                    ciudad = texto.split(':', 1)[1].strip() if ':' in texto else None
                    ciudad = self.limpiar_texto(ciudad)
        
        return vendedor, ciudad
    
    def extraer_cuentas_bancarias(self, sheet):
        """
        Extrae números de cuentas bancarias (pueden ser múltiples).
        Retorna la primera cuenta encontrada.
        """
        for fila in range(1, 50):
            for col in range(1, sheet.max_column + 1):
                texto = str(sheet.cell(fila, col).value or '').strip().upper()
                
                if 'CUENTA' in texto:
                    # Buscar número de 10 dígitos
                    for c in range(col, min(col+6, sheet.max_column+1)):
                        val = str(sheet.cell(fila, c).value or '').strip()
                        match = re.search(r'(\d{10,15})', val)
                        if match:
                            return match.group(1)
        return None
    
    def extraer_cotizacion_detalle(self, sheet):
        """
        Extrae detalles de cotización: LLANTAS, AROS, LUBRICANTES, BATERIAS.
        Retorna como string concatenado.
        """
        detalles = {}
        
        productos = ['LLANTAS', 'AROS', 'LUBRICANTES', 'BATERIAS', 'BATERÍAS']
        
        for producto in productos:
            for fila in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    texto = str(sheet.cell(fila, col).value or '').strip().upper()
                    
                    if producto in texto and 'AÑOS' not in texto:
                        # Buscar valor monetario
                        for c in range(col, min(col+4, sheet.max_column+1)):
                            val = str(sheet.cell(fila, c).value or '').strip()
                            match = re.search(r'[\$]?\s*(\d+[,\.]?\d*)', val)
                            if match and producto not in detalles:
                                detalles[producto] = match.group(0)
                                break
                        break
        
        if detalles:
            return ', '.join([f"{k}: {v}" for k, v in detalles.items()])
        return None
    
    def extraer_proveedores(self, sheet):
        """
        Extrae lista de proveedores (empresas) en la sección PROVEEDORES.
        """
        proveedores = []
        en_seccion = False
        
        for fila in range(1, sheet.max_row + 1):
            for col in range(1, min(6, sheet.max_column + 1)):
                texto = str(sheet.cell(fila, col).value or '').strip().upper()
                
                if 'PROVEEDORES' in texto or 'EMPRESA:' in texto:
                    en_seccion = True
                    col_empresa = col
                    
                    # Leer empresas debajo
                    for f in range(fila + 1, min(fila + 15, sheet.max_row + 1)):
                        empresa = str(sheet.cell(f, col_empresa).value or '').strip()
                        if empresa and len(empresa) > 2:
                            # Verificar que no sea etiqueta
                            if not any(x in empresa.upper() for x in ['OBSERVA', 'APROBADO', 'NEGADO', 'IESS', 'SRI', 'AÑO', 'CUPO']):
                                proveedores.append(empresa)
                        elif not empresa and proveedores:
                            # Si encuentra vacío y ya tiene proveedores, salir
                            break
                    
                    if proveedores:
                        return ', '.join(proveedores)
        
        return None
    
    def extraer_funcion_judicial(self, sheet, tipo):
        """
        Extrae función judicial (SI/NO o descripción).
        """
        etiqueta = f'FUNCION JUDICIAL {tipo}'
        
        for fila in range(1, 50):
            for col in range(1, sheet.max_column + 1):
                texto = str(sheet.cell(fila, col).value or '').strip().upper()
                
                if etiqueta.upper() in texto:
                    # Buscar descripción en celdas siguientes
                    for c in range(col, min(col+8, sheet.max_column+1)):
                        val = str(sheet.cell(fila, c).value or '').strip()
                        if val and len(val) > 5:
                            if 'NO REFLEJA' in val.upper() or 'NO REGISTRA' in val.upper():
                                return 'NO'
                            elif 'SI' in val.upper() or 'REFLEJA' in val.upper() or 'PENDIENTE' in val.upper():
                                return f'SI - {val}'
                            else:
                                return val
        return None
    
    def extraer_iess_sri(self, sheet, campo):
        """
        Extrae IESS o SRI: debe retornar SI/NO TIENE/descripción.
        """
        for fila in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                texto = str(sheet.cell(fila, col).value or '').strip().upper()
                
                if texto == campo:
                    # Buscar valor
                    for c in range(col+1, min(col+4, sheet.max_column+1)):
                        val = str(sheet.cell(fila, c).value or '').strip()
                        if val:
                            val_upper = val.upper()
                            if 'N/T' in val_upper or 'NO' in val_upper:
                                return 'NO TIENE'
                            elif 'SI' in val_upper or 'ACTIVO' in val_upper:
                                return 'SI'
                            else:
                                return val
        return None
    
    def extraer_archivo(self, archivo):
        """
        Extrae todos los datos de un archivo - retorna UN SOLO diccionario.
        """
        print(f" {archivo.name}")
        
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            sheet = wb.active
            
            # UN REGISTRO (una fila)
            reg = {}
            
            reg['archivo_origen'] = archivo.name
            
            # CODIGO UNICO
            reg['CODIGO_UNICO'] = self.extraer_codigo_unico(sheet)
            
            # CALIFICACION (con amarillo)
            reg['CALIFICACION'] = self.extraer_calificacion(sheet)
            
            # CARPETA COMPLETA (con amarillo)
            reg['CARPETA_COMPLETA'] = self.extraer_carpeta_completa(sheet)
            
            # DATOS PERSONALES
            reg['NOMBRE'] = self.buscar_valor_simple(sheet, ['NOMBRE'], tipo_dato='texto')
            reg['CI_TITULAR'] = self.buscar_valor_simple(sheet, ['CI: TITULAR', 'CI TITULAR'], tipo_dato='cedula')
            reg['CI_CONYUGUE'] = self.buscar_valor_simple(sheet, ['CI: CONYUGUE', 'CI: CÓNYUGE'], tipo_dato='cedula')
            reg['CI_GARANTE'] = self.extraer_ci_garante(sheet)
            reg['EDAD'] = self.extraer_edad(sheet)
            reg['ESTADO_CIVIL'] = self.buscar_valor_simple(sheet, ['ESTADO CIVIL'], tipo_dato='texto')
            
            # RUC Y AÑO
            ruc, anio = self.extraer_ruc_y_anio(sheet)
            reg['RUC'] = ruc
            reg['ANIO_RUC'] = anio
            
            # SCORES
            reg['SCORE_TITULAR'] = self.buscar_valor_simple(sheet, ['SCORE TITULAR'], tipo_dato='texto')
            reg['SCORE_CONYUGUE'] = self.buscar_valor_simple(sheet, ['SCORE CÓNYUGUE', 'SCORE CONYUGUE'], tipo_dato='texto')
            reg['SCORE_GARANTE'] = self.extraer_score_garante(sheet)
            
            # GARANTIAS
            reg['GARANTIA'] = self.buscar_valor_simple(sheet, ['GARANTIA:', 'GARANTÍA:'], tipo_dato='texto')
            reg['FIRMA_CON'] = self.buscar_valor_simple(sheet, ['FIRMA CON CÓNYUGUE:', 'FIRMA CON:'], tipo_dato='texto')
            reg['GARANTE'] = self.extraer_garante_si_no(sheet)
            reg['CONTRATO_PROV'] = self.buscar_valor_simple(sheet, ['CONTRATO DE PROV:'], tipo_dato='texto')
            reg['MATRICULA_VEHICULO'] = self.extraer_matricula_vehiculo(sheet)
            reg['COPIA_PAGOS_PREDIALES'] = self.buscar_valor_simple(sheet, ['COPIA PAGOS PREDIALES'], tipo_dato='texto')
            
            # JUDICIAL
            reg['FUNCION_JUDICIAL_TITULAR'] = self.extraer_funcion_judicial(sheet, 'TITULAR')
            reg['FUNCION_JUDICIAL_CONYUGUE'] = self.extraer_funcion_judicial(sheet, 'CÓNYUGUE')
            
            # BANCARIO
            reg['BANCO'] = self.buscar_valor_simple(sheet, ['BANCO'], tipo_dato='texto')
            reg['CUENTA'] = self.extraer_cuentas_bancarias(sheet)
            reg['CUPO'] = self.extraer_cupo(sheet)
            reg['CLIENTE_DESDE'] = self.extraer_cliente_desde(sheet)
            
            # ESTADO CUENTA
            reg['VENCIDA'] = self.buscar_valor_simple(sheet, ['VENCIDA:'], tipo_dato='numero')
            reg['POR_VENCER'] = self.extraer_por_vencer(sheet)
            reg['DOCUMENTADO'] = self.buscar_valor_simple(sheet, ['DOCUMENTADO'], tipo_dato='numero')
            
            # RIESGOS
            reg['RIESGO_TOTAL'] = self.buscar_valor_simple(sheet, ['RIESGO TOTAL'])
            reg['RIESGO_TOTAL_MAS_ALTO'] = self.buscar_valor_simple(sheet, ['RIESGO TOTAL MAS ALTO', 'RIESGO TOTAL MÁS ALTO'])
            reg['RIESGO_TOTAL_ACTUAL'] = self.buscar_valor_simple(sheet, ['RIESGO TOTAL ACTUAL'])
            
            # COTIZACION
            reg['COTIZACION'] = self.buscar_valor_simple(sheet, ['COTIZACIÓN:', 'COTIZACION:'])
            reg['COTIZACION_DETALLE'] = self.extraer_cotizacion_detalle(sheet)
            
            # VENDEDOR Y CIUDAD
            vendedor, ciudad = self.extraer_vendedor_ciudad(sheet)
            reg['VENDEDOR'] = vendedor
            reg['CIUDAD'] = ciudad
            
            # PROVEEDORES
            reg['PROVEEDORES'] = self.extraer_proveedores(sheet)
            
            # IESS, SRI
            reg['IESS'] = self.extraer_iess_sri(sheet, 'IESS')
            reg['SRI'] = self.extraer_iess_sri(sheet, 'SRI')
            
            # OBSERVACIONES
            reg['OBSERVACION'] = self.buscar_valor_simple(sheet, ['OBSERVACIÓN'])
            reg['OBSERVACION_CREDITO'] = self.buscar_valor_simple(sheet, ['OBSERVACION CREDITO'])
            reg['APROBADO_POR'] = self.buscar_valor_simple(sheet, ['APROBADO POR'])
            reg['NEGADO_POR'] = self.buscar_valor_simple(sheet, ['NEGADO POR'])
            
            # Mostrar campos importantes
            print(f"    {reg.get('NOMBRE', 'N/A')[:25]} | EST_CIVIL: {reg.get('ESTADO_CIVIL', 'N/A')} | VENDEDOR: {reg.get('VENDEDOR', 'N/A')}")
            
            wb.close()
            return reg
            
        except Exception as e:
            print(f"    ERROR: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def procesar_carpeta(self):
        """Procesa todos los archivos."""
        archivos = list(self.carpeta.glob('*.xlsx')) + list(self.carpeta.glob('*.xls'))
        archivos = [f for f in archivos if not f.name.startswith('~') and 'DATOS_LIMPIOS' not in f.name]
        
        print(f"\n {len(archivos)} archivos encontrados")
        print("=" * 80)
        
        registros = []
        for archivo in archivos:
            reg = self.extraer_archivo(archivo)
            if reg:
                registros.append(reg)
        
        print("=" * 80)
        print(f" {len(registros)} registros extraídos")
        
        return registros
    
    def exportar_excel(self, registros, ruta_salida):
        """Exporta a Excel."""
        if not registros:
            print(" Sin datos")
            return None
        
        df = pd.DataFrame(registros)
        
        # Verificar duplicados
        duplicados = df['archivo_origen'].value_counts()
        if duplicados.max() > 1:
            print(f"\n  Archivos duplicados encontrados - eliminando...")
            df = df.drop_duplicates(subset=['archivo_origen'], keep='first')
        
        # Ordenar columnas
        columnas_orden = [
            'archivo_origen', 'CODIGO_UNICO', 'CARPETA_COMPLETA', 'NOMBRE', 'CI_TITULAR',
            'CALIFICACION', 'EDAD', 'ESTADO_CIVIL', 'RUC', 'ANIO_RUC',
            'SCORE_TITULAR', 'CI_CONYUGUE', 'SCORE_CONYUGUE', 'CI_GARANTE', 'SCORE_GARANTE',
            'VENDEDOR', 'CIUDAD', 'CUPO', 'CLIENTE_DESDE', 'RIESGO_TOTAL', 'RIESGO_TOTAL_ACTUAL',
            'RIESGO_TOTAL_MAS_ALTO', 'BANCO', 'CUENTA', 'GARANTIA', 'FIRMA_CON', 'GARANTE',
            'CONTRATO_PROV', 'MATRICULA_VEHICULO', 'COPIA_PAGOS_PREDIALES',
            'FUNCION_JUDICIAL_TITULAR', 'FUNCION_JUDICIAL_CONYUGUE',
            'VENCIDA', 'POR_VENCER', 'DOCUMENTADO', 'COTIZACION', 'COTIZACION_DETALLE',
            'PROVEEDORES', 'IESS', 'SRI', 'OBSERVACION', 'OBSERVACION_CREDITO',
            'APROBADO_POR', 'NEGADO_POR'
        ]
        
        columnas_finales = [c for c in columnas_orden if c in df.columns]
        columnas_finales += [c for c in df.columns if c not in columnas_finales]
        df = df[columnas_finales]
        
        # Exportar
        df.to_excel(ruta_salida, index=False, engine='openpyxl')
        
        print(f"\nEXPORTADO: {ruta_salida}")
        print(f" {len(df)} FILAS × {len(df.columns)} COLUMNAS")
        
        # Estadísticas de completitud
        print(f"\n COMPLETITUD DE CAMPOS:")
        print("-" * 80)
        completitud = df.notna().sum().sort_values(ascending=False)
        
        for campo, count in completitud.head(20).items():
            if campo != 'archivo_origen':
                porc = (count / len(df)) * 100
                barra = "" * int(porc / 5) + "░" * (20 - int(porc / 5))
                print(f"  {campo[:30]:<30} {barra} {porc:5.1f}% ({count:2}/{len(df)})")
        
        return df


def extraer_formularios(carpeta_origen, ruta_salida):
    """Función principal."""
    print("\n" + "=" * 80)
    print(" EXTRACTOR DE FORMULARIOS EXCEL CON DETECCIÓN DE COLORES")
    print("=" * 80)
    
    if not os.path.exists(carpeta_origen):
        print(f" Carpeta no existe: {carpeta_origen}")
        return None
    
    extractor = ExtractorFormulariosCompleto(carpeta_origen)
    registros = extractor.procesar_carpeta()
    
    if registros:
        return extractor.exportar_excel(registros, ruta_salida)
    else:
        print("⚠️  No se extrajeron datos")
        return None


if __name__ == "__main__":
    carpeta_origen = r"C:\\Users\\User\\OneDrive - UNIANDES\\SEMESTRES\\NIVEL 8\\Actividades\\Pro\\copia2"
    carpeta_destino = r"C:\\Users\\User\\OneDrive - UNIANDES\\SEMESTRES\\NIVEL 8\\Actividades\\SEM1\\Lector\\ExtractorD"
    archivo_salida = os.path.join(carpeta_destino, 'DATOS_LIMPIOS_EMPROSERrVIS.xlsx')
    
    print(f" Carpeta origen: {carpeta_origen}")
    print(f"Archivo salida: {archivo_salida}")
    
    df = extraer_formularios(carpeta_origen, archivo_salida)
    
    if df is not None:
        print("\n" + "=" * 80)
        print(" PROCESO COMPLETADO EXITOSAMENTE")
        print("=" * 80)